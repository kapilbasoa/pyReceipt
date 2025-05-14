import datetime
import openpyxl
import os

class Company:
    def __init__(self, name, address, email, contact):
        self.name = name
        self.address = address
        self.email = email
        self.contact = contact

class Customer:
    def __init__(self, name, email, phone_number, address):
        self.name = name
        self.email = email
        self.phone_number = phone_number
        self.address = address

class Receipt:
    def __init__(self, company, customer, amount_paid, payment_method):
        self.company = company
        self.customer = customer
        self.amount_paid = amount_paid
        self.payment_method = payment_method
        self.receipt_number = self.generate_receipt_number()
        self.date = self.get_current_date()
        self.items = []

    def generate_receipt_number(self):
        current_time = datetime.datetime.now()
        return current_time.strftime("%Y%m%d%H%M%S")

    def get_current_date(self):
        return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def add_item(self, item_name, item_price, quantity=1):
        self.items.append({'name': item_name, 'price': item_price, 'quantity': quantity})

    def calculate_total(self):
        total = sum(item['price'] * item['quantity'] for item in self.items)
        return total

    def print_receipt(self):
        print("\nPayment Receipt")
        print("---------------------------")
        print(f"Receipt Number: {self.receipt_number}")
        print(f"Date: {self.date}")
        print(f"Customer Name: {self.customer.name}")
        print(f"Customer Email: {self.customer.email}")
        print(f"Customer Phone Number: {self.customer.phone_number}")
        print(f"Customer Address: {self.customer.address}")
        print("---------------------------")
        print(f"Company Name: {self.company.name}")
        print(f"Company Address: {self.company.address}")
        print(f"Company Email: {self.company.email}")
        print(f"Company Contact: {self.company.contact}")
        print("---------------------------")
        print("Items Purchased:")
        for item in self.items:
            item_total = item['price'] * item['quantity']
            print(f"{item['name']} x{item['quantity']} (₹{item['price']:.2f} each) = ₹{item_total:.2f}")
        print("---------------------------")
        total = self.calculate_total()
        print(f"Subtotal: ₹{total:.2f}")
        print(f"Tax (7%): ₹{total * 0.07:.2f}")
        print(f"Total Amount: ₹{total + (total * 0.07):.2f}")
        print(f"Payment Method: {self.payment_method}")
        print("---------------------------")

    def save_receipt_to_excel(self, file_name, delete_previous=False, open_after_save=False):
        if delete_previous:
            # Delete the file if it exists
            if os.path.exists(file_name):
                os.remove(file_name)

        try:
            # Try to open the existing workbook
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
        except FileNotFoundError:
            # If the file doesn't exist, create a new workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Add headers if the sheet is empty
            if sheet.max_row == 1:
                sheet['A1'] = 'Receipt Number'
                sheet['B1'] = 'Date'
                sheet['C1'] = 'Customer Name'
                sheet['D1'] = 'Customer Email'
                sheet['E1'] = 'Customer Phone Number'
                sheet['F1'] = 'Customer Address'
                sheet['G1'] = 'Company Name'
                sheet['H1'] = 'Company Address'
                sheet['I1'] = 'Company Email'
                sheet['J1'] = 'Company Contact'
                sheet['K1'] = 'Item Name'
                sheet['L1'] = 'Item Price'
                sheet['M1'] = 'Quantity'
                sheet['N1'] = 'Subtotal'
                sheet['O1'] = 'Tax (7%)'
                sheet['P1'] = 'Total Amount'
                sheet['Q1'] = 'Payment Method'

                # Apply text wrapping to header cells
                for cell in sheet['A1:Q1'][0]:
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

        # Data
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=self.receipt_number).alignment = openpyxl.styles.Alignment(wrap_text=True)
        sheet.cell(row=row, column=2, value=self.date).alignment = openpyxl.styles.Alignment(wrap_text=True)
        sheet.cell(row=row, column=3, value=self.customer.name).alignment = openpyxl.styles.Alignment(wrap_text=True)
        sheet.cell(row=row, column=4, value=self.customer.email).alignment = openpyxl.styles.Alignment(wrap_text=True)
        sheet.cell(row=row, column=5, value=self.customer.phone_number).alignment = openpyxl.styles.Alignment(wrap_text=True)
        sheet.cell(row=row, column=6, value=self.customer.address).alignment = openpyxl.styles.Alignment(wrap_text=True)
        sheet.cell(row=row, column=7, value=self.company.name).alignment = openpyxl.styles.Alignment(wrap_text=True)
        sheet.cell(row=row, column=8, value=self.company.address).alignment = openpyxl.styles.Alignment(wrap_text=True)
        sheet.cell(row=row, column=9, value=self.company.email).alignment = openpyxl.styles.Alignment(wrap_text=True)
        sheet.cell(row=row, column=10, value=self.company.contact).alignment = openpyxl.styles.Alignment(wrap_text=True)
        
        for item in self.items:
            sheet.cell(row=row, column=11, value=item['name']).alignment = openpyxl.styles.Alignment(wrap_text=True)
            sheet.cell(row=row, column=12, value=item['price']).alignment = openpyxl.styles.Alignment(wrap_text=True)
            sheet.cell(row=row, column=13, value=item['quantity']).alignment = openpyxl.styles.Alignment(wrap_text=True)
            sheet.cell(row=row, column=14, value=item['price'] * item['quantity']).alignment = openpyxl.styles.Alignment(wrap_text=True)
            row += 1

        # Totals
        total = self.calculate_total()
        sheet.cell(row=row, column=14, value=total).alignment = openpyxl.styles.Alignment(wrap_text=True)
        sheet.cell(row=row, column=15, value=total * 0.07).alignment = openpyxl.styles.Alignment(wrap_text=True)
        sheet.cell(row=row, column=16, value=total + (total * 0.07)).alignment = openpyxl.styles.Alignment(wrap_text=True)
        sheet.cell(row=row, column=17, value=self.payment_method).alignment = openpyxl.styles.Alignment(wrap_text=True)

        workbook.save(file_name)
        print(f"Receipt appended to {file_name}")
        if open_after_save:
            try:
                os.system(f"start excel {file_name}")
            except Exception as e:
                print(f"Error opening excel file: {e}")

# Example usage:
company_name = "XYZ Company"
company_address = "123 Main St, City"
company_email = "contact@xyzcompany.com"
company_contact = "123-456-7890"

customer_name = input("Enter customer name: ")
customer_email = input("Enter customer email: ")
customer_phone = int(input("Enter customer phone number: "))
customer_address = input("Enter customer address: ")
payment_method = input("Enter payment method: ")

company = Company(company_name, company_address, company_email, company_contact)
customer = Customer(customer_name, customer_email, customer_phone, customer_address)

receipt = Receipt(company, customer, 0, payment_method)

while True:
    item_name = input("Enter item name (or type 'done' to finish): ")
    if item_name.lower() == 'done':
        break
    item_price = float(input(f"Enter price for {item_name}: "))
    item_quantity = int(input(f"Enter quantity for {item_name    }: "))
    receipt.add_item(item_name, item_price, item_quantity)

# Print and save the receipt without deleting previous receipts
excel_file_name_append = "receipt_company.xlsx"
receipt.print_receipt()
receipt.save_receipt_to_excel(excel_file_name_append)

# Print and save the receipt while deleting previous receipts
excel_file_name_delete_previous = "receipt_customer.xlsx"
receipt.save_receipt_to_excel(excel_file_name_delete_previous, delete_previous=True, open_after_save=True)